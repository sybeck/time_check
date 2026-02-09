import os
import re
import time
import json
import argparse
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Dict, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeoutError

load_dotenv()
KST = timezone(timedelta(hours=9))

MAIN_SELECTOR = "#business-insights-layout__contents__main"


def must_env(key: str) -> str:
    v = os.getenv(key)
    if not v:
        raise RuntimeError(f"{key} 환경변수가 필요합니다. .env를 확인하세요.")
    return v


# ✅ 변경: 기본 날짜를 "오늘(KST)"로
def kst_today_ymd() -> str:
    return datetime.now(KST).date().strftime("%Y-%m-%d")


def build_sales_url(date_ymd: str) -> str:
    tpl = must_env("COUPANG_SALES_URL_TEMPLATE")
    return tpl.format(date=date_ymd)


def save_debug(page, prefix: str) -> None:
    os.makedirs("debug", exist_ok=True)
    page.screenshot(path=f"debug/{prefix}.png", full_page=True)
    with open(f"debug/{prefix}.html", "w", encoding="utf-8") as f:
        f.write(page.content())


# ✅ 더 빠른 대기: networkidle 의존 제거
def wait_quick(page, ms: int = 0) -> None:
    """
    - domcontentloaded는 goto에서 이미 처리하는 경우가 많음
    - networkidle은 SPA에서 오래 걸리는 경우가 많아서 제거
    - 짧은 안정화 sleep만 옵션으로 둠(기본 0ms)
    """
    if ms > 0:
        time.sleep(ms / 1000)


def normalize_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "")
    nums = re.findall(r"-?\d+", s)
    return int(nums[0]) if nums else 0


def login_coupang(page) -> None:
    login_url = must_env("COUPANG_LOGIN_URL")
    user = must_env("COUPANG_ID")
    pw = must_env("COUPANG_PW")

    page.goto(login_url, wait_until="domcontentloaded")

    scopes = [page] + list(page.frames)
    submitted = False

    for scope in scopes:
        try:
            id_loc = scope.locator(
                "input[type='email'], input[name='username'], input#username, input[name='id'], input[type='text']"
            )
            pw_loc = scope.locator("input[type='password'], input[name='password'], input#password")

            if id_loc.count() == 0 or pw_loc.count() == 0:
                continue

            id_loc.first.fill(user)
            pw_loc.first.fill(pw)

            btn = scope.locator(
                "button:has-text('로그인'), button:has-text('Login'), button[type='submit'], input[type='submit'], [role='button']:has-text('로그인')"
            )
            if btn.count() > 0:
                btn.first.click()
            else:
                pw_loc.first.press("Enter")

            submitted = True
            break
        except Exception:
            continue

    if not submitted:
        save_debug(page, "coupang_login_form_not_found")
        raise RuntimeError("로그인 폼/버튼을 찾지 못했습니다. debug/coupang_login_form_not_found.* 확인")

    # ✅ 로그인 제출 후: 긴 대기 대신 "짧은 대기 + 메인 셀렉터 확인"으로 넘기는 게 빠름
    time.sleep(3)
    wait_quick(page, int(os.getenv("POST_LOGIN_WAIT_MS", "250")))


def open_sales_url_with_retry(page, url: str, retries: int = 1) -> None:
    attempts = 0
    while True:
        attempts += 1

        # ✅ 쿠팡은 SPA라 domcontentloaded면 충분한 경우가 많음
        page.goto(url, wait_until="domcontentloaded")

        try:
            # ✅ 핵심: "페이지가 준비되었는지"는 networkidle이 아니라 MAIN_SELECTOR로 판단
            page.wait_for_selector(MAIN_SELECTOR, timeout=15_000)
            return
        except PwTimeoutError:
            if attempts > (1 + retries):
                raise RuntimeError("sales-analysis 페이지가 정상 로드되지 않았습니다(리다이렉트/2FA/로딩 문제 가능).")
            # ✅ 재시도 전 아주 짧게만
            wait_quick(page, 200)
            continue


def download_product_excel_via_dropdown(page, download_dir: str) -> str:
    """
    1) '엑셀 다운로드' 드롭다운 트리거 클릭
    2) 메뉴에서 '상품별 엑셀 다운로드' 클릭 (다운로드 발생)
    """
    os.makedirs(download_dir, exist_ok=True)

    # MAIN_SELECTOR는 open_sales_url_with_retry에서 이미 기다렸지만, 안전빵으로 짧게 한 번 더
    page.wait_for_selector(MAIN_SELECTOR, timeout=15_000)

    trigger = page.get_by_text("엑셀 다운로드", exact=True).first
    trigger.click(timeout=15_000, force=True)

    menu_item = page.get_by_text("상품별 엑셀 다운로드", exact=True).first
    menu_item.wait_for(state="visible", timeout=15_000)

    with page.expect_download(timeout=60_000) as d:
        menu_item.click(timeout=15_000, force=True)

    download = d.value
    path = os.path.join(download_dir, download.suggested_filename)
    download.save_as(path)
    return path


@dataclass
class ProductAgg:
    sales: int = 0
    qty: int = 0


def aggregate_from_excel(path: str) -> Tuple[Dict[str, ProductAgg], int, int]:
    """
    엑셀:
      - C열: 상품명
      - O열: 총 매출
      - P열: 총 판매수
      - Q열: 취소 금액 (음수)
      - R열: 취소 수량 (음수)

    집계:
      - 순매출 = O + Q
      - 순수량 = P + R
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    agg: Dict[str, ProductAgg] = {}
    total_sales = 0
    total_qty = 0

    c1, o1, p1 = ws["C1"].value, ws["O1"].value, ws["P1"].value
    start_row = 2 if any(isinstance(x, str) for x in (c1, o1, p1)) else 1

    for r in range(start_row, ws.max_row + 1):
        name = ws[f"C{r}"].value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue

        sales_o = normalize_int(ws[f"O{r}"].value)
        sales_q = normalize_int(ws[f"Q{r}"].value)
        qty_p = normalize_int(ws[f"P{r}"].value)
        qty_r = normalize_int(ws[f"R{r}"].value)

        net_sales = sales_o + sales_q
        net_qty = qty_p + qty_r

        if name not in agg:
            agg[name] = ProductAgg()
        agg[name].sales += net_sales
        agg[name].qty += net_qty

        total_sales += net_sales
        total_qty += net_qty

    return agg, total_sales, total_qty


def aggregate_by_brand(product_agg: Dict[str, ProductAgg]) -> Dict[str, ProductAgg]:
    """
    제품명을 3가지로만 분류:
      - 부담제로
      - 빠디
      - 기질 젤리
    """
    brand_map = {
        "부담제로": ["부담", "부담제로"],
        "빠디": ["빠디"],
        "기질 젤리": ["기질", "젤리", "뉴턴", "뉴턴젤리"],
    }

    result = {k: ProductAgg() for k in brand_map.keys()}

    for name, agg in product_agg.items():
        for brand, keywords in brand_map.items():
            if any(k in name for k in keywords):
                result[brand].sales += agg.sales
                result[brand].qty += agg.qty
                break

    return result


def main():
    parser = argparse.ArgumentParser(description="Coupang: product excel download -> net sales/qty -> brand summary")
    parser.add_argument("--date", help="집계 날짜 (YYYY-MM-DD). 기본: 오늘(KST)", default=None)
    parser.add_argument("--json", action="store_true", help="러너용: 마지막 줄에 JSON 1줄 출력")
    args = parser.parse_args()

    headless = os.getenv("HEADLESS", "false").lower() == "true"

    if args.date:
        ymd = datetime.strptime(args.date, "%Y-%m-%d").date().strftime("%Y-%m-%d")
    else:
        # ✅ 변경: 기본값 오늘
        ymd = kst_today_ymd()

    url = build_sales_url(ymd)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        try:
            login_coupang(page)

            open_sales_url_with_retry(page, url, retries=1)

            # ============================
            # ✅ 요청 반영(딱 이 부분만 변경)
            # - 엑셀 다운로드 실패 시: sales url 다시 로드 후 1회 재시도
            # ============================
            try:
                excel_path = download_product_excel_via_dropdown(page, download_dir="downloads")
            except Exception as e1:
                # 1) 페이지를 다시 로드
                wait_quick(page, 150)
                open_sales_url_with_retry(page, url, retries=1)

                # 2) 엑셀 다운로드 1회 재시도
                try:
                    excel_path = download_product_excel_via_dropdown(page, download_dir="downloads")
                except Exception as e2:
                    raise RuntimeError(f"엑셀 다운로드 재시도까지 실패: first={e1} / second={e2}") from e2
            # ============================

            product_agg, total_sales, total_qty = aggregate_from_excel(excel_path)
            brand_agg = aggregate_by_brand(product_agg)

            brand_summary = {brand: {"sales": v.sales, "qty": v.qty} for brand, v in brand_agg.items()}

            mapped = {
                "burdenzero": {
                    "sales": int(brand_agg.get("부담제로", ProductAgg()).sales),
                    "orders": int(brand_agg.get("부담제로", ProductAgg()).qty),
                },
                "brainology": {
                    "sales": int(brand_agg.get("기질 젤리", ProductAgg()).sales),
                    "orders": int(brand_agg.get("기질 젤리", ProductAgg()).qty),
                },
                "ppadi": {
                    "sales": int(brand_agg.get("빠디", ProductAgg()).sales),
                    "orders": int(brand_agg.get("빠디", ProductAgg()).qty),
                },
            }

            payload = {
                "status": "ok",
                "source": "coupang",
                "date": ymd,
                "target_url": url,
                "excel_path": excel_path,
                "total_sales": int(total_sales),
                "total_qty": int(total_qty),
                "brand_summary": brand_summary,
                "mapped": mapped,
            }

            if args.json:
                print(json.dumps(payload, ensure_ascii=False))
            else:
                print(payload)

        except Exception as e:
            save_debug(page, "coupang_fail")
            raise RuntimeError(f"실패: {e} (debug/coupang_fail.* 저장됨)") from e
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
